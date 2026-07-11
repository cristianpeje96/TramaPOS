"""
TramaPos · Carga masiva de productos vía Excel.

Dos operaciones:
  - generar_plantilla(): arma un .xlsx en memoria con las columnas
    esperadas y una fila de ejemplo, para que el usuario la descargue,
    la llene y la vuelva a subir.
  - procesar_archivo(): lee el .xlsx subido y crea productos/variantes.
    Se apoya en el nombre del producto para agrupar variantes: varias
    filas con el mismo "nombre_producto" se agregan como variantes de
    UN solo producto (igual que el formulario manual de creación).
    Si el producto ya existe (mismo nombre), le agrega las variantes
    nuevas en vez de duplicarlo. Si un SKU ya existe, esa fila se omite
    y se reporta en 'errores' — nunca sobreescribe silenciosamente.
"""

import io

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.productos.models import Producto, VarianteProducto

COLUMNAS = [
    "nombre_producto",
    "descripcion",
    "unidad_medida",
    "sku",
    "codigo_barras",
    "color",
    "grosor",
    "precio_venta",
    "costo_unitario",
    "stock_actual",
    "stock_minimo",
]

FILA_EJEMPLO = [
    "Hilo Guajira",
    "Hilo grueso para tejido artesanal",
    "unidad",
    "HG-MOST-GRU",
    "7701234500019",
    "Mostaza",
    "Grueso",
    8500,
    5000,
    20,
    5,
]


def generar_plantilla() -> bytes:
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.title = "Productos"

    hoja.append(COLUMNAS)
    for celda in hoja[1]:
        celda.font = Font(bold=True)

    hoja.append(FILA_EJEMPLO)
    for celda in hoja[2]:
        celda.font = Font(italic=True, color="888888")

    # Ancho de columna razonable para que se lea sin abrir a pantalla completa
    for indice, columna in enumerate(COLUMNAS, start=1):
        hoja.column_dimensions[openpyxl.utils.get_column_letter(indice)].width = max(
            14, len(columna) + 2
        )

    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def _leer_filas(contenido: bytes) -> list[dict]:
    libro = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    hoja = libro.active

    encabezados = [str(c.value).strip() if c.value else "" for c in hoja[1]]
    filas = []
    for fila_celdas in hoja.iter_rows(min_row=2):
        valores = [c.value for c in fila_celdas]
        if all(v in (None, "") for v in valores):
            continue  # fila vacía, se ignora
        fila_dict = dict(zip(encabezados, valores))
        filas.append(fila_dict)
    return filas


async def procesar_archivo(db: AsyncSession, contenido: bytes) -> dict:
    filas = _leer_filas(contenido)

    productos_creados = 0
    variantes_creadas = 0
    errores: list[str] = []
    cache_productos: dict[str, Producto] = {}

    for numero_fila, fila in enumerate(filas, start=2):  # fila 1 = encabezado
        try:
            nombre = str(fila.get("nombre_producto") or "").strip()
            sku = str(fila.get("sku") or "").strip()
            precio = fila.get("precio_venta")

            if not nombre or not sku or precio in (None, ""):
                errores.append(
                    f"Fila {numero_fila}: nombre_producto, sku y precio_venta son obligatorios"
                )
                continue

            sku_existente = await db.execute(
                select(VarianteProducto).where(VarianteProducto.sku == sku)
            )
            if sku_existente.scalar_one_or_none() is not None:
                errores.append(f"Fila {numero_fila}: el SKU '{sku}' ya existe, se omitió")
                continue

            clave_producto = nombre.lower()
            producto = cache_productos.get(clave_producto)
            if producto is None:
                resultado = await db.execute(
                    select(Producto).where(func.lower(Producto.nombre) == clave_producto)
                )
                producto = resultado.scalar_one_or_none()
                if producto is None:
                    producto = Producto(
                        nombre=nombre,
                        descripcion=str(fila.get("descripcion") or "").strip() or None,
                        unidad_medida=str(fila.get("unidad_medida") or "unidad").strip(),
                    )
                    db.add(producto)
                    await db.flush()
                    productos_creados += 1
                cache_productos[clave_producto] = producto

            variante = VarianteProducto(
                producto_id=producto.id,
                sku=sku,
                codigo_barras=str(fila.get("codigo_barras") or "").strip() or None,
                color=str(fila.get("color") or "").strip() or None,
                grosor=str(fila.get("grosor") or "").strip() or None,
                precio_venta=float(precio),
                costo_unitario=(
                    float(fila["costo_unitario"])
                    if fila.get("costo_unitario") not in (None, "")
                    else None
                ),
                stock_actual=float(fila.get("stock_actual") or 0),
                stock_minimo=float(fila.get("stock_minimo") or 0),
            )
            db.add(variante)
            variantes_creadas += 1

        except (ValueError, TypeError) as exc:
            errores.append(f"Fila {numero_fila}: dato inválido — {exc}")

    await db.commit()
    return {
        "productos_creados": productos_creados,
        "variantes_creadas": variantes_creadas,
        "errores": errores,
    }